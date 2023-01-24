# -*- coding: utf-8 -*-
"""
Created on Sun Apr 25 22:26:07 2021

@author: AC81851
"""


import tkinter
from tkinter import filedialog
import openpyxl
import ipaddress


infile = filedialog.askopenfilename()

wb = openpyxl.load_workbook(infile)
sheet = wb.active


with open('IP-Maker-Addresses.txt', 'w+') as outfile:
    outfile.write('config firewall address\n')
    for row in range(1, sheet.max_row):
        startip = sheet.cell(row, 1).value
        startip = ipaddress.IPv4Address(startip)
        endip = sheet.cell(row, 2).value
        endip = ipaddress.IPv4Address(endip)
        
        cidr = [ipaddr for ipaddr in ipaddress.summarize_address_range(startip, endip)]
        for x in cidr:
            x = x.with_prefixlen
            #print(x)
            outfile.write('edit "'+str(x)+'"\n')
            outfile.write('set subnet '+str(x)+'\n')
            outfile.write('set comment "Zoom ISDB address"\n')
            outfile.write('next\n')
    outfile.write('end')
    outfile.close()

    
with open('IP-Maker-Group.txt', 'w+') as outfile2:
    outfile2.write('config firewall addrgrp\n')
    outfile2.write('edit "Zoom_ISDB_Group"\n')
    outfile2.write('append member')
    for row in range(1, sheet.max_row):
        startip = sheet.cell(row, 1).value
        startip = ipaddress.IPv4Address(startip)
        endip = sheet.cell(row, 2).value
        endip = ipaddress.IPv4Address(endip)
        
        cidr = [ipaddr for ipaddr in ipaddress.summarize_address_range(startip, endip)]
        for x in cidr:
            x = x.with_prefixlen
            outfile2.write(' "'+str(x)+'"')
    outfile2.write('\n')
    outfile2.write('next\n')
    outfile2.write('end\n')
    outfile2.close()
    

    


